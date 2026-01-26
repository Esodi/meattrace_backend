# Admin viewsets have been removed from this project
# This file is kept to prevent import errors

from rest_framework import viewsets, status, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from django.db import transaction
from django.utils import timezone
from django.core.cache import cache
from django.conf import settings
from django.db.models import Count, Q, Sum, Avg
from django.contrib.auth import get_user_model
from datetime import timedelta
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

logger = logging.getLogger(__name__)

# Admin viewset implementations removed
# If you need custom viewsets, implement them here without admin-specific functionality

class AdminViewSet(viewsets.ModelViewSet):
    """
    Minimal placeholder AdminViewSet to avoid import errors in projects
    that removed the original admin-specific base class.
    """
    pass
    ordering = ['-created_at']

    def get_queryset(self):
        """Override to add custom filtering"""
        queryset = super().get_queryset()

        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)

        # Filter by user role
        user_role = self.request.query_params.get('user_role')
        if user_role:
            queryset = queryset.filter(user__profile__role=user_role)

        return queryset

    @action(detail=True, methods=['post'])
    def resend(self, request, pk=None):
        """Resend notification via specified channels"""
        notification = self.get_object()
        channels = request.data.get('channels', [])

        if not channels:
            return Response(
                {'error': 'channels list is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            channel_instances = NotificationChannel.objects.filter(
                name__in=channels,
                is_active=True
            )

            for channel in channel_instances:
                from .utils.notification_service import NotificationService
                NotificationService.send_via_channel(notification, channel)

            return Response({
                'message': f'Notification resent via {len(channel_instances)} channels',
                'channels': [c.name for c in channel_instances]
            })

        except Exception as e:
            return Response(
                {'error': f'Failed to resend notification: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def bulk_send(self, request):
        """Send notifications to multiple users"""
        user_ids = request.data.get('user_ids', [])
        notification_data = request.data.get('notification', {})
        channels = request.data.get('channels', [])

        if not user_ids:
            return Response(
                {'error': 'user_ids list is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not notification_data:
            return Response(
                {'error': 'notification data is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from django.contrib.auth import get_user_model
        User = get_user_model()

        users = User.objects.filter(id__in=user_ids)
        sent_notifications = []

        for user in users:
            try:
                from .utils.notification_service import NotificationService
                notification = NotificationService.create_notification(
                    user=user,
                    channels=channels,
                    **notification_data
                )
                sent_notifications.append(notification)
            except Exception as e:
                logger.error(f"Failed to send notification to user {user.username}: {str(e)}")

        return Response({
            'message': f'Sent notifications to {len(sent_notifications)} users',
            'sent_count': len(sent_notifications),
            'failed_count': len(user_ids) - len(sent_notifications)
        })

    @action(detail=False, methods=['post'])
    def broadcast(self, request):
        """Broadcast notification to users matching filters"""
        notification_data = request.data.get('notification', {})
        user_filters = request.data.get('user_filters', {})
        channels = request.data.get('channels', [])

        if not notification_data:
            return Response(
                {'error': 'notification data is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from .utils.notification_service import NotificationService
            notifications = NotificationService.broadcast_notification(
                user_filters=user_filters,
                channels=channels,
                **notification_data
            )

            return Response({
                'message': f'Broadcast sent to {len(notifications)} users',
                'sent_count': len(notifications)
            })

        except Exception as e:
            return Response(
                {'error': f'Failed to broadcast notification: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """Get notification analytics"""
        from .utils.notification_service import NotificationService

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Notification stats
        queryset = self.get_queryset()
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)

        total_notifications = queryset.count()
        read_notifications = queryset.filter(is_read=True).count()
        dismissed_notifications = queryset.filter(is_dismissed=True).count()

        # Priority breakdown
        priority_stats = queryset.values('priority').annotate(count=Count('id'))

        # Type breakdown
        type_stats = queryset.values('notification_type').annotate(count=Count('id'))

        # Delivery analytics
        delivery_analytics = NotificationService.get_delivery_analytics(start_date, end_date)

        return Response({
            'notification_stats': {
                'total': total_notifications,
                'read': read_notifications,
                'unread': total_notifications - read_notifications,
                'dismissed': dismissed_notifications,
                'read_rate': round((read_notifications / total_notifications * 100), 2) if total_notifications > 0 else 0
            },
            'priority_breakdown': list(priority_stats),
            'type_breakdown': list(type_stats),
            'delivery_analytics': delivery_analytics,
            'timestamp': timezone.now().isoformat()
        })

    @action(detail=False, methods=['post'])
    def mark_bulk_read(self, request):
        """Mark multiple notifications as read"""
        notification_ids = request.data.get('notification_ids', [])
        user_id = request.data.get('user_id')

        if not notification_ids:
            return Response(
                {'error': 'notification_ids list is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = Notification.objects.filter(id__in=notification_ids)
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        updated_count = queryset.update(is_read=True, read_at=timezone.now())

        return Response({
            'message': f'Marked {updated_count} notifications as read',
            'updated_count': updated_count
        })

    @action(detail=False, methods=['post'])
    def bulk_dismiss(self, request):
        """Dismiss multiple notifications"""
        notification_ids = request.data.get('notification_ids', [])
        user_id = request.data.get('user_id')

        if not notification_ids:
            return Response(
                {'error': 'notification_ids list is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = Notification.objects.filter(id__in=notification_ids)
        if user_id:
            queryset = queryset.filter(user_id=user_id)

        updated_count = queryset.update(is_dismissed=True, dismissed_at=timezone.now())

        return Response({
            'message': f'Dismissed {updated_count} notifications',
            'updated_count': updated_count
        })


class AdminNotificationTemplateViewSet(AdminViewSet):
    """
    ViewSet for managing notification templates with CRUD operations and testing.
    """

    from .serializers import NotificationTemplateSerializer
    from .models import NotificationTemplate

    queryset = NotificationTemplate.objects.all()
    serializer_class = NotificationTemplateSerializer
    filterset_fields = ['template_type', 'is_active']
    search_fields = ['name', 'content', 'subject']
    ordering_fields = ['name', 'template_type', 'created_at']
    ordering = ['name']

    @action(detail=True, methods=['post'])
    def test_render(self, request, pk=None):
        """Test template rendering with sample variables"""
        template = self.get_object()
        test_vars = request.data.get('variables', {})

        try:
            rendered_subject = template.render_subject(test_vars)
            rendered_content = template.render_content(test_vars)

            return Response({
                'template_name': template.name,
                'template_type': template.template_type,
                'rendered_subject': rendered_subject,
                'rendered_content': rendered_content,
                'variables_used': template.variables
            })

        except Exception as e:
            return Response(
                {'error': f'Template rendering failed: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


class AdminNotificationChannelViewSet(AdminViewSet):
    """
    ViewSet for managing notification channels with configuration and testing.
    """

    from .serializers import NotificationChannelSerializer
    from .models import NotificationChannel

    queryset = NotificationChannel.objects.all()
    serializer_class = NotificationChannelSerializer
    filterset_fields = ['channel_type', 'is_active']
    search_fields = ['name', 'channel_type', 'provider_name']
    ordering_fields = ['name', 'channel_type', 'is_active']
    ordering = ['channel_type', 'name']

    @action(detail=True, methods=['post'])
    def test_connection(self, request, pk=None):
        """Test channel connection and configuration"""
        channel = self.get_object()

        try:
            # Test based on channel type
            if channel.channel_type == 'email':
                # Test email configuration
                from django.core.mail import send_mail
                from django.conf import settings

                test_email = request.data.get('test_email')
                if not test_email:
                    return Response(
                        {'error': 'test_email is required for email channel test'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                send_mail(
                    subject='Test Notification Channel',
                    message='This is a test message from the notification system.',
                    from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'test@example.com'),
                    recipient_list=[test_email],
                    fail_silently=False
                )

            elif channel.channel_type == 'sms':
                # Test SMS configuration
                from .utils.notification_service import NotificationService

                test_phone = request.data.get('test_phone')
                if not test_phone:
                    return Response(
                        {'error': 'test_phone is required for SMS channel test'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # Create a test notification
                from django.contrib.auth import get_user_model
                User = get_user_model()
                test_user = User.objects.filter(is_staff=True).first()
                if not test_user:
                    return Response(
                        {'error': 'No staff user found for testing'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                test_notification = Notification.objects.create(
                    user=test_user,
                    notification_type='test',
                    title='Test SMS',
                    message='This is a test SMS message.'
                )

                NotificationService._send_sms(test_notification, channel, None)

            return Response({
                'message': f'{channel.channel_type.upper()} channel test successful',
                'channel': channel.name
            })

        except Exception as e:
            return Response(
                {'error': f'Channel test failed: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get'])
    def rate_limit_status(self, request, pk=None):
        """Get rate limiting status for the channel"""
        channel = self.get_object()

        # Get current usage (simplified - in production you'd aggregate from recent deliveries)
        from django.utils import timezone
        from datetime import timedelta

        now = timezone.now()
        minute_ago = now - timedelta(minutes=1)
        hour_ago = now - timedelta(hours=1)
        day_ago = now - timedelta(days=1)

        recent_deliveries = NotificationDelivery.objects.filter(
            channel=channel,
            created_at__gte=day_ago
        )

        minute_count = recent_deliveries.filter(created_at__gte=minute_ago).count()
        hour_count = recent_deliveries.filter(created_at__gte=hour_ago).count()
        day_count = recent_deliveries.count()

        return Response({
            'channel': channel.name,
            'limits': {
                'per_minute': channel.rate_limit_per_minute,
                'per_hour': channel.rate_limit_per_hour,
                'per_day': channel.rate_limit_per_day
            },
            'current_usage': {
                'minute': minute_count,
                'hour': hour_count,
                'day': day_count
            },
            'is_limited': channel.is_rate_limited()
        })


class AdminNotificationScheduleViewSet(AdminViewSet):
    """
    ViewSet for managing notification schedules with CRUD and execution control.
    """

    from .serializers import NotificationScheduleSerializer
    from .models import NotificationSchedule

    queryset = NotificationSchedule.objects.select_related('created_by').prefetch_related('recipient_users', 'channels').all()
    serializer_class = NotificationScheduleSerializer
    filterset_fields = ['schedule_type', 'frequency', 'is_active', 'notification_type']
    search_fields = ['title', 'title_template', 'message_template']
    ordering_fields = ['title', 'scheduled_at', 'created_at']
    ordering = ['-scheduled_at']

    @action(detail=True, methods=['post'])
    def execute_now(self, request, pk=None):
        """Execute scheduled notification immediately"""
        schedule = self.get_object()

        try:
            from .utils.notification_service import NotificationService

            # Get recipients
            recipients = list(schedule.recipient_users.all())

            # Add users from groups
            if schedule.recipient_groups:
                from django.contrib.auth import get_user_model
                User = get_user_model()

                for group in schedule.recipient_groups:
                    if group == 'abbatoirs':
                        recipients.extend(User.objects.filter(profile__role='Abbatoir'))
                    elif group == 'processors':
                        recipients.extend(User.objects.filter(profile__role='Processor'))
                    elif group == 'shop_owners':
                        recipients.extend(User.objects.filter(profile__role='ShopOwner'))
                    elif group == 'admins':
                        recipients.extend(User.objects.filter(profile__role='Admin'))

            # Remove duplicates
            recipients = list(set(recipients))

            sent_count = 0
            for user in recipients:
                try:
                    NotificationService.create_notification(
                        user=user,
                        notification_type=schedule.notification_type,
                        title=schedule.title_template,
                        message=schedule.message_template,
                        template_vars=schedule.template_variables,
                        channels=list(schedule.channels.all())
                    )
                    sent_count += 1
                except Exception as e:
                    logger.error(f"Failed to send scheduled notification to {user.username}: {str(e)}")

            return Response({
                'message': f'Schedule executed successfully. Sent to {sent_count} recipients.',
                'sent_count': sent_count,
                'total_recipients': len(recipients)
            })

        except Exception as e:
            return Response(
                {'error': f'Failed to execute schedule: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def toggle_active(self, request, pk=None):
        """Toggle schedule active status"""
        schedule = self.get_object()
        schedule.is_active = not schedule.is_active
        schedule.save()

        return Response({
            'message': f'Schedule {"activated" if schedule.is_active else "deactivated"}',
            'is_active': schedule.is_active
        })

    @action(detail=False, methods=['get'])
    def due_schedules(self, request):
        """Get schedules that are due for execution"""
        from django.utils import timezone

        due_schedules = self.get_queryset().filter(
            is_active=True,
            scheduled_at__lte=timezone.now()
        )

        serializer = self.get_serializer(due_schedules, many=True)
        return Response({
            'due_schedules': serializer.data,
            'count': due_schedules.count()
        })


class AdminNotificationDeliveryViewSet(AdminViewSet):
    """
    ViewSet for monitoring notification delivery status and analytics.
    """

    from .serializers import NotificationDeliverySerializer
    from .models import NotificationDelivery

    queryset = NotificationDelivery.objects.select_related(
        'notification', 'channel', 'recipient'
    ).all()
    serializer_class = NotificationDeliverySerializer
    filterset_fields = [
        'status', 'channel', 'recipient', 'notification__notification_type'
    ]
    search_fields = ['notification__title', 'recipient__username', 'channel__name']
    ordering_fields = ['created_at', 'sent_at', 'delivered_at', 'failed_at']
    ordering = ['-created_at']

    def get_queryset(self):
        """Override to add custom filtering"""
        queryset = super().get_queryset()

        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)

        return queryset

    @action(detail=True, methods=['post'])
    def retry_delivery(self, request, pk=None):
        """Retry a failed delivery"""
        delivery = self.get_object()

        if delivery.status not in ['failed', 'retrying']:
            return Response(
                {'error': 'Only failed deliveries can be retried'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from .utils.notification_service import NotificationService
            NotificationService.send_via_channel(delivery.notification, delivery.channel)

            return Response({
                'message': 'Delivery retry initiated',
                'delivery_id': delivery.id
            })

        except Exception as e:
            return Response(
                {'error': f'Failed to retry delivery: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def bulk_retry(self, request):
        """Retry multiple failed deliveries"""
        delivery_ids = request.data.get('delivery_ids', [])

        if not delivery_ids:
            return Response(
                {'error': 'delivery_ids list is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        queryset = self.get_queryset().filter(
            id__in=delivery_ids,
            status__in=['failed', 'retrying']
        )

        retried_count = 0
        for delivery in queryset:
            try:
                from .utils.notification_service import NotificationService
                NotificationService.send_via_channel(delivery.notification, delivery.channel)
                retried_count += 1
            except Exception as e:
                logger.error(f"Failed to retry delivery {delivery.id}: {str(e)}")

        return Response({
            'message': f'Initiated retry for {retried_count} deliveries',
            'retried_count': retried_count,
            'total_requested': len(delivery_ids)
        })

    @action(detail=False, methods=['get'])
    def delivery_stats(self, request):
        """Get delivery statistics"""
        from .utils.notification_service import NotificationService

        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        analytics = NotificationService.get_delivery_analytics(start_date, end_date)

        # Add additional stats
        queryset = self.get_queryset()
        if start_date:
            queryset = queryset.filter(created_at__gte=start_date)
        if end_date:
            queryset = queryset.filter(created_at__lte=end_date)

        # Status breakdown
        status_stats = queryset.values('status').annotate(count=Count('id'))

        # Channel performance
        channel_performance = queryset.values('channel__name').annotate(
            total=Count('id'),
            successful=Count('id', filter=Q(status='delivered')),
            failed=Count('id', filter=Q(status='failed'))
        ).annotate(
            success_rate=Case(
                When(total=0, then=0),
                default=(F('successful') * 100.0 / F('total'))
            )
        )

        return Response({
            'analytics': analytics,
            'status_breakdown': list(status_stats),
            'channel_performance': list(channel_performance),
            'timestamp': timezone.now().isoformat()
        })


# ══════════════════════════════════════════════════════════════════════════════
# ADMIN DASHBOARD VIEWSETS
# ══════════════════════════════════════════════════════════════════════════════

class AdminDashboardViewSet(viewsets.ViewSet):
    """
    ViewSet for admin dashboard overview and statistics
    """
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get dashboard overview statistics"""
        from .models import (
            User, UserProfile, ProcessingUnit, Shop, Animal, Product,
            Order, Sale, SystemAlert, Activity
        )
        from .serializers import AdminDashboardStatsSerializer

        # User statistics
        total_users = User.objects.count()
        user_role_counts = UserProfile.objects.values('role').annotate(count=Count('id'))

        abbatoirs_count = next((item['count'] for item in user_role_counts if item['role'] == 'Abbatoir'), 0)
        processors_count = next((item['count'] for item in user_role_counts if item['role'] == 'Processor'), 0)
        shop_owners_count = next((item['count'] for item in user_role_counts if item['role'] == 'ShopOwner'), 0)
        admins_count = next((item['count'] for item in user_role_counts if item['role'] == 'Admin'), 0)

        # Entity statistics
        total_processing_units = ProcessingUnit.objects.count()
        active_processing_units = ProcessingUnit.objects.filter(
            members__is_active=True
        ).distinct().count()

        total_shops = Shop.objects.count()
        active_shops = Shop.objects.filter(
            members__is_active=True
        ).distinct().count()

        # Data statistics
        total_animals = Animal.objects.count()
        total_products = Product.objects.count()
        total_orders = Order.objects.count()
        total_sales = Sale.objects.count()

        # Recent activity (last 7 days)
        seven_days_ago = timezone.now() - timedelta(days=7)
        recent_animals_count = Animal.objects.filter(created_at__gte=seven_days_ago).count()
        recent_products_count = Product.objects.filter(created_at__gte=seven_days_ago).count()
        recent_orders_count = Order.objects.filter(created_at__gte=seven_days_ago).count()
        recent_activities_count = Activity.objects.filter(timestamp__gte=seven_days_ago).count()

        # System health
        system_health_status = "healthy"  # Default
        try:
            from .models import SystemHealth
            latest_health = SystemHealth.objects.filter(component='api').first()
            if latest_health:
                system_health_status = latest_health.status
        except:
            pass

        active_alerts_count = SystemAlert.objects.filter(is_active=True).count()

        stats_data = {
            'total_users': total_users,
            'total_abbatoirs': abbatoirs_count,
            'total_processors': processors_count,
            'total_shop_owners': shop_owners_count,
            'total_admins': admins_count,
            'total_processing_units': total_processing_units,
            'total_shops': total_shops,
            'active_processing_units': active_processing_units,
            'active_shops': active_shops,
            'total_animals': total_animals,
            'total_products': total_products,
            'total_orders': total_orders,
            'total_sales': total_sales,
            'recent_animals_count': recent_animals_count,
            'recent_products_count': recent_products_count,
            'recent_orders_count': recent_orders_count,
            'recent_activities_count': recent_activities_count,
            'system_health_status': system_health_status,
            'active_alerts_count': active_alerts_count,
        }

        serializer = AdminDashboardStatsSerializer(stats_data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def recent_activity(self, request):
        """Get recent system activity for dashboard"""
        from .models import Activity
        from .serializers import AdminRecentActivitySerializer

        limit = int(request.query_params.get('limit', 20))
        activities = Activity.objects.select_related('user').order_by('-timestamp')[:limit]

        activity_data = []
        for activity in activities:
            activity_data.append({
                'id': activity.id,
                'user': activity.user.username if activity.user else 'System',
                'activity_type': activity.activity_type,
                'title': activity.title,
                'description': activity.description,
                'timestamp': activity.timestamp.isoformat(),
                'entity_type': activity.entity_type,
                'entity_id': activity.entity_id,
            })

        response_data = {
            'activities': activity_data,
            'total_count': len(activity_data),
            'pagination': {
                'limit': limit,
                'has_more': Activity.objects.count() > limit
            }
        }

        serializer = AdminRecentActivitySerializer(response_data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def supply_chain_stats(self, request):
        """Get supply chain statistics for the monitor dashboard"""
        from .models import Animal, Product, ProcessingUnit, Shop, TransferRequest, UserProfile
        
        today = timezone.now().date()
        
        # Today's transfers: animals and products transferred today
        animals_transferred_today = Animal.objects.filter(
            transferred_at__date=today
        ).count()
        
        products_transferred_today = Product.objects.filter(
            transferred_at__date=today
        ).count()
        
        todays_transfers = animals_transferred_today + products_transferred_today
        
        # Active locations: abbatoirs + processing units + shops that are active
        active_abbatoirs = UserProfile.objects.filter(role='Abbatoir', user__is_active=True).count()
        active_processing_units = ProcessingUnit.objects.filter(is_active=True).count()
        active_shops = Shop.objects.filter(is_active=True).count()
        active_locations = active_abbatoirs + active_processing_units + active_shops
        
        # Pending transfer requests (transit alerts)
        pending_transfers = TransferRequest.objects.filter(status='pending').count()
        
        return Response({
            'todays_transfers': todays_transfers,
            'animals_transferred_today': animals_transferred_today,
            'products_transferred_today': products_transferred_today,
            'active_locations': active_locations,
            'active_abbatoirs': active_abbatoirs,
            'active_processing_units': active_processing_units,
            'active_shops': active_shops,
            'pending_transfers': pending_transfers,
            'date': today.isoformat()
        })
    
    @action(detail=False, methods=['get'])
    def map_locations(self, request):
        """
        Get all active locations with coordinates for the Supply Chain Map.
        Returns processing units, shops, and abbatoirs with their geographic coordinates.
        """
        from .models import ProcessingUnit, Shop, UserProfile
        
        locations = []
        
        # Get active processing units with coordinates
        processing_units = ProcessingUnit.objects.filter(
            is_active=True,
            latitude__isnull=False,
            longitude__isnull=False
        )
        for pu in processing_units:
            locations.append({
                'id': f'pu_{pu.id}',
                'name': pu.name,
                'type': 'Processing Unit',
                'lat': float(pu.latitude),
                'lng': float(pu.longitude),
                'location': pu.location or '',
                'contact_email': pu.contact_email or '',
                'contact_phone': pu.contact_phone or '',
            })
        
        # Get active shops with coordinates
        shops = Shop.objects.filter(
            is_active=True,
            latitude__isnull=False,
            longitude__isnull=False
        )
        for shop in shops:
            locations.append({
                'id': f'shop_{shop.id}',
                'name': shop.name,
                'type': 'Shop',
                'lat': float(shop.latitude),
                'lng': float(shop.longitude),
                'location': shop.location or '',
                'contact_email': shop.contact_email or '',
                'contact_phone': shop.contact_phone or '',
            })
        
        # Get abbatoirs with coordinates
        abbatoirs = UserProfile.objects.filter(
            role='Abbatoir',
            user__is_active=True,
            latitude__isnull=False,
            longitude__isnull=False
        ).select_related('user')
        for abbatoir in abbatoirs:
            locations.append({
                'id': f'abbatoir_{abbatoir.id}',
                'name': f"{abbatoir.user.first_name} {abbatoir.user.last_name}".strip() or abbatoir.user.username,
                'type': 'Abbatoir',
                'lat': float(abbatoir.latitude),
                'lng': float(abbatoir.longitude),
                'location': abbatoir.address or '',
                'contact_email': abbatoir.user.email or '',
                'contact_phone': abbatoir.phone or '',
            })
        
        # Summary statistics
        total_pu = ProcessingUnit.objects.filter(is_active=True).count()
        total_shops = Shop.objects.filter(is_active=True).count()
        total_abbatoirs = UserProfile.objects.filter(role='Abbatoir', user__is_active=True).count()
        
        geocoded_pu = processing_units.count()
        geocoded_shops = shops.count()
        geocoded_abbatoirs = abbatoirs.count()
        
        return Response({
            'locations': locations,
            'total_count': len(locations),
            'summary': {
                'processing_units': {'total': total_pu, 'geocoded': geocoded_pu},
                'shops': {'total': total_shops, 'geocoded': geocoded_shops},
                'abbatoirs': {'total': total_abbatoirs, 'geocoded': geocoded_abbatoirs},
            }
        })

class AdminUserViewSet(viewsets.ModelViewSet):
    """
    ViewSet for admin user management
    """
    permission_classes = [IsAuthenticated]
    queryset = None  # Will be set in get_queryset
    serializer_class = None  # Will be set in get_serializer_class
    filterset_fields = ['is_active', 'is_staff', 'is_superuser']
    search_fields = ['username', 'email', 'first_name', 'last_name']
    ordering_fields = ['username', 'email', 'date_joined', 'last_login']
    ordering = ['-date_joined']

    def get_queryset(self):
        """Override to add role filtering"""
        queryset = get_user_model().objects.select_related('profile').all()

        role = self.request.query_params.get('role')
        if role:
            queryset = queryset.filter(profile__role=role)

        processing_unit = self.request.query_params.get('processing_unit')
        if processing_unit:
            queryset = queryset.filter(profile__processing_unit_id=processing_unit)

        shop = self.request.query_params.get('shop')
        if shop:
            queryset = queryset.filter(profile__shop_id=shop)

        return queryset

    def get_serializer_class(self):
        from .serializers import AdminUserListSerializer, AdminUserDetailSerializer, AdminUserCreateUpdateSerializer
        if self.action == 'list':
            return AdminUserListSerializer
        elif self.action in ['retrieve']:
            return AdminUserDetailSerializer
        elif self.action in ['create', 'update', 'partial_update']:
            return AdminUserCreateUpdateSerializer
        return AdminUserListSerializer

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Activate a user account"""
        user = self.get_object()
        user.is_active = True
        user.save()
        return Response({'message': f'User {user.username} activated'})

    @action(detail=True, methods=['post'])
    def deactivate(self, request, pk=None):
        """Deactivate a user account"""
        user = self.get_object()
        user.is_active = False
        user.save()
        return Response({'message': f'User {user.username} deactivated'})

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        """Reset user password (admin action)"""
        user = self.get_object()
        new_password = request.data.get('new_password')
        if not new_password:
            return Response(
                {'error': 'new_password is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        user.set_password(new_password)
        user.save()

        # Log the password reset
        from .models import UserAuditLog
        UserAuditLog.objects.create(
            performed_by=request.user,
            affected_user=user,
            action='password_changed',
            description=f'Password reset by admin {request.user.username}'
        )

        return Response({'message': f'Password reset for user {user.username}'})


class AdminProcessingUnitViewSet(viewsets.ModelViewSet):
    """
    ViewSet for admin processing unit management
    """
    permission_classes = [IsAuthenticated]
    queryset = None  # Will be set in get_queryset
    serializer_class = None  # Will be set in get_serializer_class
    filterset_fields = ['is_active']
    search_fields = ['name', 'location', 'contact_email']
    ordering_fields = ['name', 'created_at', 'updated_at']
    ordering = ['-created_at']

    def get_queryset(self):
        from .models import ProcessingUnit
        return ProcessingUnit.objects.prefetch_related('members', 'products').all()

    def get_serializer_class(self):
        from .serializers import AdminProcessingUnitSerializer
        return AdminProcessingUnitSerializer

    @action(detail=True, methods=['get'])
    def members(self, request, pk=None):
        """Get processing unit members"""
        processing_unit = self.get_object()
        from .serializers import ProcessingUnitUserSerializer
        members = processing_unit.members.select_related('user').all()
        serializer = ProcessingUnitUserSerializer(members, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def products(self, request, pk=None):
        """Get processing unit products"""
        processing_unit = self.get_object()
        from .serializers import ProductSerializer
        products = processing_unit.products.select_related('animal', 'processing_unit').all()
        serializer = ProductSerializer(products, many=True)
        return Response(serializer.data)


class AdminShopViewSet(viewsets.ModelViewSet):
    """
    ViewSet for admin shop management
    """
    permission_classes = [IsAuthenticated]
    queryset = None  # Will be set in get_queryset
    serializer_class = None  # Will be set in get_serializer_class
    filterset_fields = ['is_active']
    search_fields = ['name', 'location', 'contact_email']
    ordering_fields = ['name', 'created_at', 'updated_at']
    ordering = ['-created_at']

    def get_queryset(self):
        from .models import Shop
        return Shop.objects.prefetch_related('members', 'inventory', 'orders', 'sales').all()

    def get_serializer_class(self):
        from .serializers import AdminShopSerializer
        return AdminShopSerializer

    @action(detail=True, methods=['get'])
    def members(self, request, pk=None):
        """Get shop members"""
        shop = self.get_object()
        from .serializers import ShopUserSerializer
        members = shop.members.select_related('user').all()
        serializer = ShopUserSerializer(members, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def inventory(self, request, pk=None):
        """Get shop inventory"""
        shop = self.get_object()
        from .serializers import InventorySerializer
        inventory = shop.inventory.select_related('product').all()
        serializer = InventorySerializer(inventory, many=True)
        return Response(serializer.data)


class AdminAnimalViewSet(viewsets.ModelViewSet):
    """
    ViewSet for admin animal management with full CRUD.
    Allows admins to add animals at any point in the traceability chain.
    """
    permission_classes = [IsAuthenticated]
    queryset = None  # Will be set in get_queryset
    serializer_class = None  # Will be set in get_serializer_class
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['species', 'slaughtered', 'transferred_to']
    search_fields = ['animal_id', 'animal_name', 'abbatoir__username']
    ordering_fields = ['created_at', 'slaughtered_at', 'transferred_at']
    ordering = ['-created_at']

    def get_serializer_class(self):
        from .serializers import AdminAnimalOverviewSerializer, AdminAnimalCreateUpdateSerializer
        if self.action in ['create', 'update', 'partial_update']:
            return AdminAnimalCreateUpdateSerializer
        return AdminAnimalOverviewSerializer

    def get_queryset(self):
        """Override to add lifecycle status filtering"""
        from .models import Animal
        queryset = Animal.objects.select_related('abbatoir', 'transferred_to').all()

        lifecycle_status = self.request.query_params.get('lifecycle_status')
        if lifecycle_status:
            if lifecycle_status == 'HEALTHY':
                queryset = queryset.filter(slaughtered=False, transferred_to__isnull=True)
            elif lifecycle_status == 'SLAUGHTERED':
                queryset = queryset.filter(slaughtered=True, transferred_to__isnull=True)
            elif lifecycle_status == 'TRANSFERRED':
                queryset = queryset.filter(transferred_to__isnull=False)
            elif lifecycle_status == 'SEMI-TRANSFERRED':
                # Animals with some parts transferred but not all
                queryset = queryset.filter(
                    transferred_to__isnull=True,
                    slaughter_parts__transferred_to__isnull=False
                ).distinct()

        return queryset


class AdminProductViewSet(viewsets.ModelViewSet):
    """
    ViewSet for admin product management with full CRUD.
    Allows admins to add products at processing units or directly to shops.
    """
    permission_classes = [IsAuthenticated]
    queryset = None  # Will be set in get_queryset
    serializer_class = None  # Will be set in get_serializer_class
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['product_type', 'transferred_to', 'received_by_shop', 'category']
    search_fields = ['name', 'batch_number', 'animal__animal_id']
    ordering_fields = ['created_at', 'transferred_at']
    ordering = ['-created_at']

    def get_serializer_class(self):
        from .serializers import AdminProductOverviewSerializer, AdminProductCreateUpdateSerializer
        if self.action in ['create', 'update', 'partial_update']:
            return AdminProductCreateUpdateSerializer
        return AdminProductOverviewSerializer

    def get_queryset(self):
        from .models import Product
        return Product.objects.select_related(
            'animal', 'processing_unit', 'transferred_to', 'received_by_shop', 'category'
        ).all()


class AdminSlaughterPartViewSet(viewsets.ModelViewSet):
    """
    ViewSet for admin slaughter part management with full CRUD.
    Allows admins to add slaughter parts from slaughtered animals.
    """
    permission_classes = [IsAuthenticated]
    queryset = None  # Will be set in get_queryset
    serializer_class = None  # Will be set in get_serializer_class
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['part_type', 'transferred_to', 'used_in_product']
    search_fields = ['part_id', 'animal__animal_id', 'animal__animal_name']
    ordering_fields = ['created_at', 'weight']
    ordering = ['-created_at']

    def get_serializer_class(self):
        from .serializers import SlaughterPartSerializer, AdminSlaughterPartCreateUpdateSerializer
        if self.action in ['create', 'update', 'partial_update']:
            return AdminSlaughterPartCreateUpdateSerializer
        return SlaughterPartSerializer

    def get_queryset(self):
        from .models import SlaughterPart
        queryset = SlaughterPart.objects.select_related('animal', 'transferred_to').all()
        
        # Filter by animal_id if provided
        animal_id = self.request.query_params.get('animal_id')
        if animal_id:
            queryset = queryset.filter(animal_id=animal_id)
        
        return queryset


class AdminAbbatoirViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet for listing abbatoirs for admin selection dropdowns.
    Read-only - abbatoirs are managed through the user management interface.
    """
    permission_classes = [IsAuthenticated]
    queryset = None  # Will be set in get_queryset
    serializer_class = None  # Will be set in get_serializer_class
    search_fields = ['username', 'email', 'first_name', 'last_name']
    ordering_fields = ['username', 'first_name', 'date_joined']
    ordering = ['username']

    def get_serializer_class(self):
        from .serializers import AdminAbbatoirListSerializer
        return AdminAbbatoirListSerializer

    def get_queryset(self):
        from django.contrib.auth import get_user_model
        User = get_user_model()
        return User.objects.filter(
            profile__role='Abbatoir',
            is_active=True
        ).select_related('profile').prefetch_related('animals')


class AdminAnalyticsViewSet(viewsets.ViewSet):
    """
    ViewSet for admin analytics and reporting.
    Analytics are cached for 5 minutes to improve performance.
    """
    permission_classes = [IsAuthenticated]

    @action(detail=False, methods=['get'])
    def overview(self, request):
        """Get analytics overview for specified period"""
        from .serializers import AdminAnalyticsSerializer
        from .models import User, Animal, Product, Order, Sale

        # Parse date parameters
        period = request.query_params.get('period', '30d')
        end_date = timezone.now().date()

        if period == '7d':
            start_date = end_date - timedelta(days=7)
        elif period == '30d':
            start_date = end_date - timedelta(days=30)
        elif period == '90d':
            start_date = end_date - timedelta(days=90)
        else:
            start_date = end_date - timedelta(days=30)

        # User metrics
        new_users_count = User.objects.filter(
            date_joined__date__gte=start_date,
            date_joined__date__lte=end_date
        ).count()

        active_users_count = User.objects.filter(
            last_login__date__gte=start_date
        ).distinct().count()

        # Entity metrics
        new_animals_count = Animal.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).count()

        new_products_count = Product.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).count()

        new_orders_count = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).count()

        new_sales_count = Sale.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        ).count()

        # Processing metrics - calculate from actual data
        # Processing efficiency: % of animals that have been processed or slaughtered
        total_animals = Animal.objects.filter(created_at__date__gte=start_date).count()
        # Animal model uses 'processed' and 'slaughtered' boolean fields, not 'status'
        processed_animals = Animal.objects.filter(
            created_at__date__gte=start_date,
            processed=True
        ).count()
        processing_efficiency = round((processed_animals / total_animals * 100), 2) if total_animals > 0 else 0

        # Transfer success rate: calculate from TransferRequest model
        try:
            from .models import TransferRequest
            total_transfers = TransferRequest.objects.filter(created_at__date__gte=start_date).count()
            successful_transfers = TransferRequest.objects.filter(
                created_at__date__gte=start_date,
                status='approved'
            ).count()
            transfer_success_rate = round((successful_transfers / total_transfers * 100), 2) if total_transfers > 0 else 0
        except Exception:
            transfer_success_rate = 0

        # Financial metrics
        total_sales_value = Sale.objects.filter(
            created_at__date__gte=start_date
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        average_order_value = Order.objects.filter(
            created_at__date__gte=start_date
        ).aggregate(avg=Avg('total_amount'))['avg'] or 0

        # System metrics - based on actual data availability
        # System uptime: placeholder (would need actual monitoring)
        system_uptime = 0  # No monitoring data available
        error_rate = 0  # No error tracking data available

        analytics_data = {
            'period': period,
            'start_date': start_date,
            'end_date': end_date,
            'new_users_count': new_users_count,
            'active_users_count': active_users_count,
            'new_animals_count': new_animals_count,
            'new_products_count': new_products_count,
            'new_orders_count': new_orders_count,
            'new_sales_count': new_sales_count,
            'processing_efficiency': processing_efficiency,
            'transfer_success_rate': transfer_success_rate,
            'total_sales_value': total_sales_value,
            'average_order_value': average_order_value,
            'system_uptime': system_uptime,
            'error_rate': error_rate,
        }

        serializer = AdminAnalyticsSerializer(analytics_data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def daily_stats(self, request):
        """Get daily statistics for the last N days"""
        days = int(request.query_params.get('days', 7))
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days)

        from .models import Animal, Product, Order, Sale, User

        daily_stats = []
        for i in range(days):
            date = start_date + timedelta(days=i)

            stats = {
                'date': date.isoformat(),
                'new_users': User.objects.filter(date_joined__date=date).count(),
                'new_animals': Animal.objects.filter(created_at__date=date).count(),
                'new_products': Product.objects.filter(created_at__date=date).count(),
                'new_orders': Order.objects.filter(created_at__date=date).count(),
                'new_sales': Sale.objects.filter(created_at__date=date).count(),
            }
            daily_stats.append(stats)

        return Response({
            'period': f'{days} days',
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'daily_stats': daily_stats
        })

    @action(detail=False, methods=['get'])
    def weekly_stats(self, request):
        """Get weekly statistics for the last N weeks"""
        weeks = int(request.query_params.get('weeks', 4))
        end_date = timezone.now().date()
        start_date = end_date - timedelta(weeks=weeks)

        from .models import Animal, Product, Order, Sale, User

        weekly_stats = []
        for i in range(weeks):
            week_start = start_date + timedelta(weeks=i)
            week_end = week_start + timedelta(days=6)

            stats = {
                'week': f'{week_start.isoformat()} to {week_end.isoformat()}',
                'new_users': User.objects.filter(
                    date_joined__date__gte=week_start,
                    date_joined__date__lte=week_end
                ).count(),
                'new_animals': Animal.objects.filter(
                    created_at__date__gte=week_start,
                    created_at__date__lte=week_end
                ).count(),
                'new_products': Product.objects.filter(
                    created_at__date__gte=week_start,
                    created_at__date__lte=week_end
                ).count(),
                'new_orders': Order.objects.filter(
                    created_at__date__gte=week_start,
                    created_at__date__lte=week_end
                ).count(),
                'new_sales': Sale.objects.filter(
                    created_at__date__gte=week_start,
                    created_at__date__lte=week_end
                ).count(),
            }
            weekly_stats.append(stats)

        return Response({
            'period': f'{weeks} weeks',
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'weekly_stats': weekly_stats
        })

    @action(detail=False, methods=['get'])
    def custom_report(self, request):
        """Get custom analytics report with granular filters"""
        from .models import Animal, Product, Order, Sale, User, ProcessingUnit, Shop
        
        # Filters
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        animal_id = request.query_params.get('animal_id')
        product_id = request.query_params.get('product_id')
        shop_id = request.query_params.get('shop_id')
        abbatoir_id = request.query_params.get('abbatoir_id')
        processing_unit_id = request.query_params.get('processing_unit_id')

        # Base filters
        q_animals = Q()
        q_products = Q()
        q_orders = Q()
        q_sales = Q()

        if start_date:
            q_animals &= Q(created_at__date__gte=start_date)
            q_products &= Q(created_at__date__gte=start_date)
            q_orders &= Q(created_at__date__gte=start_date)
            q_sales &= Q(created_at__date__gte=start_date)
        if end_date:
            q_animals &= Q(created_at__date__lte=end_date)
            q_products &= Q(created_at__date__lte=end_date)
            q_orders &= Q(created_at__date__lte=end_date)
            q_sales &= Q(created_at__date__lte=end_date)
        
        if animal_id:
            q_animals &= Q(id=animal_id)
            q_products &= Q(animal_id=animal_id)
            # Orders and Sales might not link directly to animal, skipping for simplicity or would need joins
        
        if product_id:
            q_products &= Q(id=product_id)
            q_orders &= Q(products__id=product_id)
            q_sales &= Q(products__id=product_id)

        if shop_id:
            q_orders &= Q(shop_id=shop_id)
            q_sales &= Q(shop_id=shop_id)
            # Products link to shop via received_by_shop
            q_products &= Q(received_by_shop_id=shop_id)

        if abbatoir_id:
            q_animals &= Q(abbatoir_id=abbatoir_id)
        
        if processing_unit_id:
            q_products &= Q(processing_unit_id=processing_unit_id)

        # Execute queries
        animals = Animal.objects.filter(q_animals).select_related('abbatoir', 'transferred_to')
        products = Product.objects.filter(q_products).select_related('animal', 'processing_unit', 'received_by_shop')
        orders = Order.objects.filter(q_orders).select_related('shop', 'customer')
        sales = Sale.objects.filter(q_sales).select_related('shop', 'seller')

        data = {
            'summary': {
                'total_animals': animals.count(),
                'total_products': products.count(),
                'total_orders': orders.count(),
                'total_sales': sales.count(),
                'total_revenue': sales.aggregate(total=Sum('total_amount'))['total'] or 0,
            },
            'filters': {
                'start_date': start_date,
                'end_date': end_date,
                'animal_id': animal_id,
                'product_id': product_id,
                'shop_id': shop_id,
                'abbatoir_id': abbatoir_id,
                'processing_unit_id': processing_unit_id,
            }
        }

        # Serializing list data for the table
        data['animals'] = [{
            'id': a.id,
            'animal_id': a.animal_id,
            'species': a.species,
            'abbatoir': a.abbatoir.username if a.abbatoir else 'Unknown',
            'created_at': a.created_at.isoformat(),
            'slaughtered': a.slaughtered
        } for a in animals[:100]] # Limit to 100 for API response

        data['products'] = [{
            'id': p.id,
            'name': p.name,
            'batch_number': p.batch_number,
            'processing_unit': p.processing_unit.name if p.processing_unit else 'N/A',
            'created_at': p.created_at.isoformat()
        } for p in products[:100]]

        return Response(data)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """Export filtered report to Excel"""
        from .models import Animal, Product, Order, Sale
        
        # Reuse same filter logic (ideally would refactor this into a method)
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        shop_id = request.query_params.get('shop_id')
        processing_unit_id = request.query_params.get('processing_unit_id')
        
        q_animals = Q()
        q_products = Q()
        q_orders = Q()
        q_sales = Q()

        if start_date:
            q_animals &= Q(created_at__date__gte=start_date)
            q_products &= Q(created_at__date__gte=start_date)
            q_orders &= Q(created_at__date__gte=start_date)
            q_sales &= Q(created_at__date__gte=start_date)
        if end_date:
            q_animals &= Q(created_at__date__lte=end_date)
            q_products &= Q(created_at__date__lte=end_date)
            q_orders &= Q(created_at__date__lte=end_date)
            q_sales &= Q(created_at__date__lte=end_date)
        
        if shop_id:
            q_orders &= Q(shop_id=shop_id)
            q_sales &= Q(shop_id=shop_id)
        if processing_unit_id:
            q_products &= Q(processing_unit_id=processing_unit_id)

        # Create Workbook
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "Executive Summary"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        summary_data = [
            ["MeatTrace Industry Report", ""],
            ["Generated At", timezone.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Period", f"{start_date or 'Beginning'} to {end_date or 'Now'}"],
            ["", ""],
            ["Metric", "Value"],
            ["Total Animals Registered", Animal.objects.filter(q_animals).count()],
            ["Total Products Created", Product.objects.filter(q_products).count()],
            ["Total Orders Placed", Order.objects.filter(q_orders).count()],
            ["Total Sales Recorded", Sale.objects.filter(q_sales).count()],
            ["Total Revenue", Sale.objects.filter(q_sales).aggregate(total=Sum('total_amount'))['total'] or 0],
        ]
        
        for row in summary_data:
            ws_summary.append(row)
            
        # Style Summary
        for cell in ws_summary["A5:B5"][0]:
            cell.font = header_font
            cell.fill = header_fill
        
        # Animals Sheet
        ws_animals = wb.create_sheet("Animals")
        headers = ["ID", "Animal ID", "Species", "Abbatoir", "Registered At", "Slaughtered"]
        ws_animals.append(headers)
        for cell in ws_animals[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        for a in Animal.objects.filter(q_animals).select_related('abbatoir')[:1000]:
            ws_animals.append([
                a.id, a.animal_id, a.species, 
                a.abbatoir.username if a.abbatoir else 'Unknown',
                a.created_at.strftime("%Y-%m-%d %H:%M"),
                "Yes" if a.slaughtered else "No"
            ])
            
        # Products Sheet
        ws_products = wb.create_sheet("Products")
        headers = ["ID", "Name", "Batch", "Processing Unit", "Created At"]
        ws_products.append(headers)
        for cell in ws_products[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        for p in Product.objects.filter(q_products).select_related('processing_unit')[:1000]:
            ws_products.append([
                p.id, p.name, p.batch_number,
                p.processing_unit.name if p.processing_unit else 'N/A',
                p.created_at.strftime("%Y-%m-%d %H:%M")
            ])

        # Prepare response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"meattrace_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response


# Import missing serializer for inventory
class InventorySerializer(serializers.ModelSerializer):
    """Serializer for inventory items"""
    product_name = serializers.CharField(source='product.name', read_only=True)
    product_type = serializers.CharField(source='product.product_type', read_only=True)

    class Meta:
        from .models import Inventory
        model = Inventory
        fields = '__all__'

# ══════════════════════════════════════════════════════════════════════════════
# SYSTEM CONFIGURATION VIEWSETS
# ══════════════════════════════════════════════════════════════════════════════

from .models import SystemConfiguration, FeatureFlag
from .serializers import FeatureFlagSerializer

class SystemConfigurationViewSet(viewsets.ModelViewSet):
    """ViewSet for system configuration management."""
    queryset = SystemConfiguration.objects.all()
    permission_classes = [IsAdminUser]
    
    class SystemConfigSerializer(serializers.ModelSerializer):
        class Meta:
            model = SystemConfiguration
            fields = '__all__'
    
    serializer_class = SystemConfigSerializer


class FeatureFlagViewSet(viewsets.ModelViewSet):
    """ViewSet for feature flag management."""
    queryset = FeatureFlag.objects.all()
    permission_classes = [IsAdminUser]
    serializer_class = FeatureFlagSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['is_enabled', 'environment']
    search_fields = ['name', 'description']


# ══════════════════════════════════════════════════════════════════════════════
# GOVERNMENT ADMIN VIEWSETS
# ══════════════════════════════════════════════════════════════════════════════

from .models import ComplianceAudit, Certification, RegistrationApplication, ApprovalWorkflow
from .serializers import (
    ComplianceAuditSerializer, CertificationSerializer,
    RegistrationApplicationSerializer, ApprovalWorkflowSerializer
)
from .throttling import AdminRateThrottle

class AdminComplianceAuditViewSet(viewsets.ModelViewSet):
    queryset = ComplianceAudit.objects.all().order_by('-created_at')
    serializer_class = ComplianceAuditSerializer
    permission_classes = [IsAdminUser]
    throttle_classes = [AdminRateThrottle]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    # Only include actual model fields - status and audit_type exist on the model
    filterset_fields = ['status', 'audit_type']
    search_fields = ['processing_unit__name', 'shop__name', 'abbatoir__username', 'auditor__username']

class AdminCertificationViewSet(viewsets.ModelViewSet):
    queryset = Certification.objects.all().order_by('-created_at')
    serializer_class = CertificationSerializer
    permission_classes = [IsAdminUser]
    throttle_classes = [AdminRateThrottle]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status', 'cert_type']
    search_fields = ['name', 'certificate_number', 'issuing_authority']

class RegistrationApplicationViewSet(viewsets.ModelViewSet):
    queryset = RegistrationApplication.objects.all().order_by('-created_at')
    serializer_class = RegistrationApplicationSerializer
    permission_classes = [IsAdminUser]
    throttle_classes = [AdminRateThrottle]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['status', 'entity_type']
    search_fields = ['entity_name', 'business_license_number', 'tin_number']

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        application = self.get_object()
        application.status = 'approved'
        application.decision_date = timezone.now()
        application.reviewed_by = request.user
        application.save()
        
        # Logic to actually creating the entity (ProcessingUnit/Shop) typically goes here
        # For now we just mark approved.
        
        return Response({'status': 'approved'})

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        application = self.get_object()
        application.status = 'rejected'
        application.decision_date = timezone.now()
        application.reviewed_by = request.user
        application.review_notes = request.data.get('reason', '')
        application.save()
        return Response({'status': 'rejected'})

class ApprovalWorkflowViewSet(viewsets.ModelViewSet):
    queryset = ApprovalWorkflow.objects.all()
    serializer_class = ApprovalWorkflowSerializer
    permission_classes = [IsAdminUser]
    throttle_classes = [AdminRateThrottle]
